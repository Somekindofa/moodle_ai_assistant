"""
Export all video elicitation annotations to JSON + Excel and push both
files to OwnCloud via WebDAV.

Destination:
  https://cloud.minesparis.psl.eu/remote.php/dav/files/<uuid>/
      craftpilot_shared/exports/
          annotations_export.json
          annotations_export.xlsx

Run manually:
  /root/miniconda3/envs/moodle_backend/bin/python /opt/craftpilot_backend/export_to_owncloud.py

Or via cron (see /etc/cron.d/craftpilot-export).
"""

import json
import logging
import os
import sys
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pymysql
import pymysql.cursors
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────

DB = dict(host="localhost", user="moodleuser", password=os.getenv("MOODLE_DB_PASSWORD", ""), database="moodle")

def _get_webdav_config():
    """Read all WebDAV settings from Moodle DB (single source of truth).

    The canonical values live in mdl_config_plugins where plugin='local_videoelicit'.
    Renaming the shared folder on OwnCloud requires only updating webdav_shared_folder there.
    """
    conn = pymysql.connect(**DB, cursorclass=pymysql.cursors.DictCursor)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT name, value FROM mdl_config_plugins "
                "WHERE plugin='local_videoelicit' AND name IN "
                "('webdav_base_url','webdav_username','webdav_password','webdav_user_id','webdav_shared_folder')"
            )
            cfg = {row["name"]: row["value"] for row in cur.fetchall()}
    finally:
        conn.close()
    return cfg


def _build_export_dir(cfg):
    base     = cfg.get("webdav_base_url", "https://cloud.minesparis.psl.eu")
    user_id  = cfg["webdav_user_id"]
    folder   = cfg["webdav_shared_folder"].strip("/")
    return f"{base}/remote.php/dav/files/{user_id}/{folder}/exports"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("/tmp/craftpilot_export.log"),
    ],
)
log = logging.getLogger(__name__)

# ── Database query ────────────────────────────────────────────────────────────

QUERY = """
    SELECT
        a.id,
        a.craft,
        a.task,
        a.starttime,
        a.endtime,
        a.transcription,
        a.transcriptionstatus,
        a.reviewstatus,
        a.judgestatus,
        a.taggingstatus,
        a.issalient,
        a.tags,
        a.reviewresults,
        a.judgedecision,
        a.timecreated,
        a.timemodified,
        v.filename  AS video_filename,
        v.source_type,
        v.external_url AS video_url,
        u.username,
        u.firstname,
        u.lastname,
        u.email
    FROM mdl_local_videoelicit_annotations a
    LEFT JOIN mdl_local_videoelicit_videos v ON v.id = a.videoid
    LEFT JOIN mdl_user u ON u.id = a.userid
    ORDER BY a.timecreated DESC
"""


def fetch_annotations():
    conn = pymysql.connect(**DB, cursorclass=pymysql.cursors.DictCursor)
    try:
        with conn.cursor() as cur:
            cur.execute(QUERY)
            rows = cur.fetchall()
    finally:
        conn.close()

    for row in rows:
        # Parse JSON columns
        for col in ("tags", "reviewresults", "judgedecision"):
            raw = row.get(col)
            if raw:
                try:
                    row[col] = json.loads(raw)
                except (json.JSONDecodeError, TypeError):
                    pass
        # Human-readable timestamps
        for col in ("timecreated", "timemodified"):
            if row.get(col):
                row[col] = datetime.fromtimestamp(row[col]).isoformat()

    return rows


# ── JSON export ───────────────────────────────────────────────────────────────

class _Encoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


def build_json(annotations):
    payload = {
        "exported_at": datetime.now().isoformat(),
        "total": len(annotations),
        "annotations": annotations,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, cls=_Encoder).encode("utf-8")


# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EEF2FF")

COLUMNS = [
    ("id",                   "id",                  8),
    ("craft",                "craft",               14),
    ("task",                 "task",                20),
    ("video_filename",       "video_filename",      30),
    ("start_time_s",         "starttime",           10),
    ("end_time_s",           "endtime",             10),
    ("username",             "username",            18),
    ("first_name",           "firstname",           14),
    ("last_name",            "lastname",            14),
    ("transcription",        "transcription",       60),
    ("tags",                 "_tags_flat",          40),
    ("transcription_status", "transcriptionstatus", 16),
    ("review_status",        "reviewstatus",        14),
    ("judge_status",         "judgestatus",         14),
    ("tagging_status",       "taggingstatus",       14),
    ("is_salient",           "issalient",           10),
    ("created_at",           "timecreated",         20),
    ("judge_decision",       "_judgedecision_flat", 40),
    ("video_url",            "video_url",           50),
]


def _flat_tags(tags):
    if not isinstance(tags, list):
        return ""
    return "; ".join(f"{t.get('category','?')}:{t.get('name','?')}" for t in tags)


def _flat_judge(jd):
    if isinstance(jd, dict):
        return json.dumps(jd, ensure_ascii=False)
    if isinstance(jd, str):
        return jd
    return ""


def build_excel(annotations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Annotations"

    # Header row
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, ann in enumerate(annotations, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        ann["_tags_flat"] = _flat_tags(ann.get("tags"))
        ann["_judgedecision_flat"] = _flat_judge(ann.get("judgedecision"))

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            val = ann.get(key)
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 60

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── WebDAV upload ─────────────────────────────────────────────────────────────

def webdav_put(export_dir, username, password, path, data, content_type):
    url = f"{export_dir}/{path}"
    auth = (username, password)

    # Ensure the exports/ directory exists (MKCOL is idempotent on existing dirs)
    requests.request("MKCOL", f"{export_dir}/", auth=auth, timeout=30)

    resp = requests.put(
        url,
        data=data,
        auth=auth,
        headers={"Content-Type": content_type},
        timeout=60,
    )
    if resp.status_code not in (200, 201, 204):
        raise RuntimeError(f"WebDAV PUT failed: {resp.status_code} {resp.text[:200]}")
    log.info("Uploaded %s → %s (%d bytes)", path, url, len(data))


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("Reading WebDAV config from Moodle DB…")
    cfg = _get_webdav_config()
    export_dir = _build_export_dir(cfg)
    username   = cfg["webdav_username"]
    password   = cfg["webdav_password"]
    log.info("  Export target: %s", export_dir)

    log.info("Fetching annotations from MySQL…")
    annotations = fetch_annotations()
    log.info("  %d annotations found", len(annotations))

    log.info("Building JSON export…")
    json_bytes = build_json(annotations)

    log.info("Building Excel export…")
    xlsx_bytes = build_excel(annotations)

    log.info("Uploading to OwnCloud…")
    webdav_put(export_dir, username, password, "annotations_export.json",  json_bytes, "application/json")
    webdav_put(export_dir, username, password, "annotations_export.xlsx",  xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    log.info("Done. %d annotations exported at %s", len(annotations), datetime.now().isoformat())


if __name__ == "__main__":
    main()
